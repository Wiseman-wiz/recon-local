from pprint import pprint
import openpyxl
from bson.objectid import ObjectId
import pymongo


client = pymongo.MongoClient(host='localhost', port=27017)


class Upload_Checker():
    def __init__(self, filename: str, company_code: str):
        self.filename = filename
        self.uploaded_filepath = f'/home/admin/apps/bank_recon/development/recon/uploadfiles/AJE/{filename}'
        self.wb = openpyxl.load_workbook(self.uploaded_filepath)
        self.ws = self.wb["Form"]
        self.max_col = self.ws.max_column
        self.max_row = self.ws.max_row
        self.pass_list = []
        self.fail_list = []
        self.key_ref = self.get_dict(3)
        self.company_code = company_code

    def get_dict(self, row_num):
        my_dict = {}
        for i in range(1, self.max_col + 1):
            cell_obj = self.ws.cell(row = row_num, column = i)
            my_dict[i] = cell_obj.value
        return my_dict

    def get_pass_fail_list(self):
        db = client[f'{self.company_code}_AJE_records']
        pb_draft = db['pb_aje_draft']
        gl_draft = db["gl_aje_draft"]

        current_row = 4
        while(current_row <= self.max_row):
            my_dict = {
                self.key_ref[i].lower() : self.ws.cell(row = current_row, column = i).value
                for i in range(1, self.max_col + 1)
            }
            if not all(value == None for value in my_dict.values()):
                self.pass_list.append(my_dict) if my_dict['recon'] == 1 else self.fail_list.append(my_dict)
            current_row+=1
        
        for entry in self.pass_list:
            transaction_aje_number = entry["transaction_aje_number"]
            stringed_id = transaction_aje_number.split("-")[1]
            original_id = ObjectId(stringed_id)
            data_filter = {"original_id" : original_id}
            (pb_draft.update_one(
                data_filter,
                {"$set" : entry})
            if "PB" in self.filename
            else gl_draft.update_one(
                data_filter,
                {"$set" : entry}))

        return {
            "pass_list" : self.pass_list,
            "fail_list" : self.fail_list
        }

### test run below
# filename = 'PB_AJE_7kcIaLE9.xlsx'
# filename = 'GL_AJE_1qCWLyhf.xlsx'
# file_path_argument = "/home/admin/apps/bank_recon/development/recon/app/controllers/helpers/aje_export_module/output/GL_AJE_GUiaPGl6.xlsx"
# up_check = Upload_Checker(file_path_argument)
# res = up_check.get_pass_fail_list()
# pprint(res)



    